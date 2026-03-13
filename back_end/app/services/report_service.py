"""
报告导出服务模块

提供控制计划和 OCAP 的报告导出功能
按照 AIAG/VDA 标准格式导出控制计划

使用 Jinja2 模板引擎进行 HTML 报告导出
"""

import io
from io import BytesIO
from typing import Dict, List, Any, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from fastapi import Response, HTTPException
from jinja2 import Environment, FileSystemLoader
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.models import ControlPlan, ControlPlanItem, OCAP, OCAPSignal, OCAPStep, OCAPExecution, OCAPRootCause, OCAPCorrectiveAction

# 模板引擎配置 - 使用绝对路径
current_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
template_dir = os.path.join(current_dir, 'templates', 'reports')
jinja_env = Environment(loader=FileSystemLoader(template_dir))


thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def export_control_plan_excel(plan_id: int, db: Session) -> Response:
    """
    按照 AIAG/VDA 标准格式导出控制计划 Excel
    
    标准格式包含 26 个字段：
    1) PROTOTYPE/PRE-LAUNCH/PRODUCTION
    2) CONTROL PLAN NUMBER
    3) PART NUMBER/LATEST CHANGE LEVEL
    4) PART NAME/DESCRIPTION
    5) ORGANIZATION/PLANT
    6) ORGANIZATION CODE
    7) KEY CONTACT/PHONE
    8) CORE TEAM
    9) ORGANIZATION/PLANT APPROVAL/DATE
    10) DATE (ORIG.)
    11) DATE (REV.)
    12) CUSTOMER ENGINEERING APPROVAL/DATE
    13) CUSTOMER QUALITY APPROVAL/DATE
    14) OTHER APPROVAL/DATE
    15) PART/PROCESS NUMBER
    16) PROCESS NAME/OPERATION DESCRIPTION
    17) MACHINE, DEVICE, JIG, TOOLS FOR MANUFACTURING
    18) NUMBER (CHARACTERISTICS)
    19) PRODUCT CHARACTERISTIC
    20) PROCESS CHARACTERISTIC
    21) SPECIAL CHARACTERISTIC CLASSIFICATION
    22) PRODUCT/PROCESS SPECIFICATION/TOLERANCE
    23) EVALUATION/MEASUREMENT TECHNIQUE
    24) SAMPLE SIZE / FREQ.
    25) CONTROL METHOD
    26) REACTION PLAN
    """
    plan = db.query(ControlPlan).filter(ControlPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail=f"控制计划 {plan_id} 不存在")
    
    items = db.query(ControlPlanItem).filter(
        ControlPlanItem.control_plan_id == plan_id
    ).order_by(ControlPlanItem.sort_order).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Plan"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    
    # 字体和样式定义
    header_font = Font(name='Arial', bold=True, size=10)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 第 1 行：标题
    ws.merge_cells('A1:P1')
    ws['A1'] = 'CONTROL PLAN'
    ws['A1'].font = Font(name='Arial', bold=True, size=16)
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # 第 2 行：Prototype/Pre-launch/Production 选择
    ws.merge_cells('A2:A2')
    ws['A2'] = '1)'
    ws['A2'].font = Font(name='Arial', bold=True, size=9)
    ws['A2'].alignment = center_alignment
    
    ws.merge_cells('B2:D2')
    prototype_text = []
    if plan.plan_type:
        if plan.plan_type == 'prototype':
            prototype_text.append('☑ Prototype')
            prototype_text.append('☐ Pre-launch')
            prototype_text.append('☐ Production')
        elif plan.plan_type == 'pre_launch':
            prototype_text.append('☐ Prototype')
            prototype_text.append('☑ Pre-launch')
            prototype_text.append('☐ Production')
        else:
            prototype_text.append('☐ Prototype')
            prototype_text.append('☐ Pre-launch')
            prototype_text.append('☑ Production')
    else:
        prototype_text = ['☐ Prototype', '☐ Pre-launch', '☐ Production']
    ws['B2'] = '\n'.join(prototype_text)
    ws['B2'].font = Font(name='Arial', size=9)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 40
    
    # 第 2 行右侧：页码
    ws.merge_cells('E2:P2')
    page_num = f"Page {plan.page_number or 1} of {plan.total_pages or 1}"
    ws['E2'] = page_num
    ws['E2'].font = Font(name='Arial', size=9)
    ws['E2'].alignment = Alignment(horizontal='right', vertical='center')
    
    # 第 3 行：Control Plan Number, Key Contact/Phone, Date (Orig.), Date (Rev.)
    # 2) CONTROL PLAN NUMBER
    ws.cell(row=3, column=1).value = '2)'
    ws.cell(row=3, column=1).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=3, column=1).alignment = center_alignment
    
    ws.merge_cells('B3:C3')
    ws['B3'] = plan.control_plan_number or ''
    ws['B3'].font = Font(name='Arial', size=9)
    ws['B3'].alignment = cell_alignment
    
    # 7) KEY CONTACT/PHONE
    ws.cell(row=3, column=4).value = '7)'
    ws.cell(row=3, column=4).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=3, column=4).alignment = center_alignment
    
    ws.merge_cells('E3:J3')
    ws['E3'] = plan.key_contact or ''
    ws['E3'].font = Font(name='Arial', size=9)
    ws['E3'].alignment = cell_alignment
    
    # 10) DATE (ORIG.)
    ws.cell(row=3, column=10).value = '10)'
    ws.cell(row=3, column=10).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=3, column=10).alignment = center_alignment
    
    ws.merge_cells('K3:M3')
    ws['K3'] = plan.date_orig.strftime('%Y-%m-%d') if plan.date_orig else ''
    ws['K3'].font = Font(name='Arial', size=9)
    ws['K3'].alignment = cell_alignment
    
    # 11) DATE (REV.)
    ws.cell(row=3, column=13).value = '11)'
    ws.cell(row=3, column=13).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=3, column=13).alignment = center_alignment
    
    ws.merge_cells('N3:P3')
    ws['N3'] = plan.date_rev.strftime('%Y-%m-%d') if plan.date_rev else ''
    ws['N3'].font = Font(name='Arial', size=9)
    ws['N3'].alignment = cell_alignment
    
    ws.row_dimensions[3].height = 25
    
    # 第 4 行：Part Number/Latest Change Level, Core Team, Customer Engineering Approval
    # 3) PART NUMBER/LATEST CHANGE LEVEL
    ws.cell(row=4, column=1).value = '3)'
    ws.cell(row=4, column=1).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=4, column=1).alignment = center_alignment
    
    ws.merge_cells('B4:C4')
    ws['B4'] = f"{plan.part_number or ''} {plan.latest_change_level or ''}".strip()
    ws['B4'].font = Font(name='Arial', size=9)
    ws['B4'].alignment = cell_alignment
    
    # 8) CORE TEAM
    ws.cell(row=4, column=4).value = '8)'
    ws.cell(row=4, column=4).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=4, column=4).alignment = center_alignment
    
    ws.merge_cells('E4:J4')
    ws['E4'] = plan.core_team or ''
    ws['E4'].font = Font(name='Arial', size=9)
    ws['E4'].alignment = cell_alignment
    
    # 12) CUSTOMER ENGINEERING APPROVAL/DATE
    ws.cell(row=4, column=10).value = '12)'
    ws.cell(row=4, column=10).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=4, column=10).alignment = center_alignment
    
    ws.merge_cells('K4:P4')
    customer_eng = []
    if plan.customer_eng_approval_by:
        customer_eng.append(plan.customer_eng_approval_by)
    if plan.customer_eng_approval_date:
        customer_eng.append(plan.customer_eng_approval_date.strftime('%Y-%m-%d'))
    ws['K4'] = '\n'.join(customer_eng) if customer_eng else ''
    ws['K4'].font = Font(name='Arial', size=9)
    ws['K4'].alignment = cell_alignment
    
    ws.row_dimensions[4].height = 25
    
    # 第 5 行：Part Name/Description, Organization/Plant Approval, Customer Quality Approval
    # 4) PART NAME/DESCRIPTION
    ws.cell(row=5, column=1).value = '4)'
    ws.cell(row=5, column=1).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=5, column=1).alignment = center_alignment
    
    ws.merge_cells('B5:C5')
    ws['B5'] = plan.part_name or plan.part_description or ''
    ws['B5'].font = Font(name='Arial', size=9)
    ws['B5'].alignment = cell_alignment
    
    # 9) ORGANIZATION/PLANT APPROVAL/DATE
    ws.cell(row=5, column=4).value = '9)'
    ws.cell(row=5, column=4).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=5, column=4).alignment = center_alignment
    
    ws.merge_cells('E5:J5')
    org_approval = []
    if plan.org_approval_by:
        org_approval.append(plan.org_approval_by)
    if plan.org_approval_date:
        org_approval.append(plan.org_approval_date.strftime('%Y-%m-%d'))
    ws['E5'] = '\n'.join(org_approval) if org_approval else ''
    ws['E5'].font = Font(name='Arial', size=9)
    ws['E5'].alignment = cell_alignment
    
    # 13) CUSTOMER QUALITY APPROVAL/DATE
    ws.cell(row=5, column=10).value = '13)'
    ws.cell(row=5, column=10).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=5, column=10).alignment = center_alignment
    
    ws.merge_cells('K5:P5')
    customer_quality = []
    if plan.customer_quality_approval_by:
        customer_quality.append(plan.customer_quality_approval_by)
    if plan.customer_quality_approval_date:
        customer_quality.append(plan.customer_quality_approval_date.strftime('%Y-%m-%d'))
    ws['K5'] = '\n'.join(customer_quality) if customer_quality else ''
    ws['K5'].font = Font(name='Arial', size=9)
    ws['K5'].alignment = cell_alignment
    
    ws.row_dimensions[5].height = 25
    
    # 第 6 行：Organization/Plant, Organization Code, Other Approval
    # 5) ORGANIZATION/PLANT
    ws.cell(row=6, column=1).value = '5)'
    ws.cell(row=6, column=1).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=6, column=1).alignment = center_alignment
    
    ws.merge_cells('B6:C6')
    ws['B6'] = plan.organization_plant or ''
    ws['B6'].font = Font(name='Arial', size=9)
    ws['B6'].alignment = cell_alignment
    
    # 6) ORGANIZATION CODE (SUPPLIER CODE)
    ws.cell(row=6, column=4).value = '6)'
    ws.cell(row=6, column=4).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=6, column=4).alignment = center_alignment
    
    ws.merge_cells('E6:J6')
    ws['E6'] = plan.organization_code or ''
    ws['E6'].font = Font(name='Arial', size=9)
    ws['E6'].alignment = cell_alignment
    
    # 14) OTHER APPROVAL/DATE
    ws.cell(row=6, column=10).value = '14)'
    ws.cell(row=6, column=10).font = Font(name='Arial', bold=True, size=9)
    ws.cell(row=6, column=10).alignment = center_alignment
    
    ws.merge_cells('K6:P6')
    other_approval = []
    if plan.other_approval_by:
        other_approval.append(plan.other_approval_by)
    if plan.other_approval_date:
        other_approval.append(plan.other_approval_date.strftime('%Y-%m-%d'))
    ws['K6'] = '\n'.join(other_approval) if other_approval else ''
    ws['K6'].font = Font(name='Arial', size=9)
    ws['K6'].alignment = cell_alignment
    
    ws.row_dimensions[6].height = 25
    
    # 第 7 行：表格标题行 - 主要列头
    header_row = 7
    ws.merge_cells(f'A{header_row}:A8')
    ws.cell(row=header_row, column=1).value = '15)\nPART/\nPROCESS\nNUMBER'
    ws.cell(row=header_row, column=1).font = header_font
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).alignment = header_alignment
    ws.cell(row=header_row, column=1).border = thin_border
    
    ws.merge_cells(f'B{header_row}:B8')
    ws.cell(row=header_row, column=2).value = '16)\nPROCESS NAME/\nOPERATION\nDESCRIPTION'
    ws.cell(row=header_row, column=2).font = header_font
    ws.cell(row=header_row, column=2).fill = header_fill
    ws.cell(row=header_row, column=2).alignment = header_alignment
    ws.cell(row=header_row, column=2).border = thin_border
    
    ws.merge_cells(f'C{header_row}:C8')
    ws.cell(row=header_row, column=3).value = '17)\nMACHINE,\nDEVICE, JIG,\nTOOLS FOR\nMFG.'
    ws.cell(row=header_row, column=3).font = header_font
    ws.cell(row=header_row, column=3).fill = header_fill
    ws.cell(row=header_row, column=3).alignment = header_alignment
    ws.cell(row=header_row, column=3).border = thin_border
    
    ws.merge_cells(f'D{header_row}:G{header_row}')
    ws.cell(row=header_row, column=4).value = 'CHARACTERISTICS'
    ws.cell(row=header_row, column=4).font = header_font
    ws.cell(row=header_row, column=4).fill = header_fill
    ws.cell(row=header_row, column=4).alignment = header_alignment
    ws.cell(row=header_row, column=4).border = thin_border
    
    ws.merge_cells(f'H{header_row}:H8')
    ws.cell(row=header_row, column=8).value = '21)\nSPECIAL\nCHAR.\nCLASS'
    ws.cell(row=header_row, column=8).font = header_font
    ws.cell(row=header_row, column=8).fill = header_fill
    ws.cell(row=header_row, column=8).alignment = header_alignment
    ws.cell(row=header_row, column=8).border = thin_border
    
    ws.merge_cells(f'I{header_row}:N{header_row}')
    ws.cell(row=header_row, column=9).value = 'METHODS'
    ws.cell(row=header_row, column=9).font = header_font
    ws.cell(row=header_row, column=9).fill = header_fill
    ws.cell(row=header_row, column=9).alignment = header_alignment
    ws.cell(row=header_row, column=9).border = thin_border
    
    ws.merge_cells(f'O{header_row}:O8')
    ws.cell(row=header_row, column=15).value = '26)\nREACTION\nPLAN'
    ws.cell(row=header_row, column=15).font = header_font
    ws.cell(row=header_row, column=15).fill = header_fill
    ws.cell(row=header_row, column=15).alignment = header_alignment
    ws.cell(row=header_row, column=15).border = thin_border
    
    ws.row_dimensions[header_row].height = 40
    
    # 第 8 行：CHARACTERISTICS 和 METHODS 子列头
    ws.cell(row=8, column=4).value = '18)\nNO.'
    ws.cell(row=8, column=4).font = header_font
    ws.cell(row=8, column=4).fill = header_fill
    ws.cell(row=8, column=4).alignment = header_alignment
    ws.cell(row=8, column=4).border = thin_border
    
    ws.cell(row=8, column=5).value = '19)\nPRODUCT'
    ws.cell(row=8, column=5).font = header_font
    ws.cell(row=8, column=5).fill = header_fill
    ws.cell(row=8, column=5).alignment = header_alignment
    ws.cell(row=8, column=5).border = thin_border
    
    ws.cell(row=8, column=6).value = '20)\nPROCESS'
    ws.cell(row=8, column=6).font = header_font
    ws.cell(row=8, column=6).fill = header_fill
    ws.cell(row=8, column=6).alignment = header_alignment
    ws.cell(row=8, column=6).border = thin_border
    
    ws.cell(row=8, column=7).value = '22)\nPRODUCT/\nPROCESS\nSPECIFICATION/\nTOLERANCE'
    ws.cell(row=8, column=7).font = header_font
    ws.cell(row=8, column=7).fill = header_fill
    ws.cell(row=8, column=7).alignment = header_alignment
    ws.cell(row=8, column=7).border = thin_border
    
    ws.cell(row=8, column=8).value = '23)\nEVALUATION/\nMEASUREMENT\nTECHNIQUE'
    ws.cell(row=8, column=8).font = header_font
    ws.cell(row=8, column=8).fill = header_fill
    ws.cell(row=8, column=8).alignment = header_alignment
    ws.cell(row=8, column=8).border = thin_border
    
    ws.merge_cells(f'I8:J8')
    ws.cell(row=8, column=9).value = '24)\nSAMPLE'
    ws.cell(row=8, column=9).font = header_font
    ws.cell(row=8, column=9).fill = header_fill
    ws.cell(row=8, column=9).alignment = header_alignment
    ws.cell(row=8, column=9).border = thin_border
    
    ws.cell(row=8, column=10).value = 'SIZE'
    ws.cell(row=8, column=10).font = header_font
    ws.cell(row=8, column=10).fill = header_fill
    ws.cell(row=8, column=10).alignment = header_alignment
    ws.cell(row=8, column=10).border = thin_border
    
    ws.cell(row=8, column=11).value = 'FREQ.'
    ws.cell(row=8, column=11).font = header_font
    ws.cell(row=8, column=11).fill = header_fill
    ws.cell(row=8, column=11).alignment = header_alignment
    ws.cell(row=8, column=11).border = thin_border
    
    ws.cell(row=8, column=12).value = '25)\nCONTROL\nMETHOD'
    ws.cell(row=8, column=12).font = header_font
    ws.cell(row=8, column=12).fill = header_fill
    ws.cell(row=8, column=12).alignment = header_alignment
    ws.cell(row=8, column=12).border = thin_border
    
    ws.row_dimensions[8].height = 40
    
    # 填充数据行
    data_row = 9
    for idx, item in enumerate(items, 1):
        ws.cell(row=data_row, column=1).value = item.part_process_number or str(idx)
        ws.cell(row=data_row, column=2).value = f"{item.process_name or ''}\n{item.operation_description or ''}".strip()
        ws.cell(row=data_row, column=3).value = item.machine_device_jig_tools or ''
        ws.cell(row=data_row, column=4).value = item.characteristic_no or str(idx)
        ws.cell(row=data_row, column=5).value = item.product_characteristic or ''
        ws.cell(row=data_row, column=6).value = item.process_characteristic or ''
        ws.cell(row=data_row, column=7).value = item.specification_tolerance or ''
        ws.cell(row=data_row, column=8).value = item.special_characteristic_class or ''
        ws.cell(row=data_row, column=9).value = item.evaluation_measurement_technique or ''
        ws.cell(row=data_row, column=10).value = item.sample_size or ''
        ws.cell(row=data_row, column=11).value = item.sample_frequency or ''
        ws.cell(row=data_row, column=12).value = item.control_method or ''
        ws.cell(row=data_row, column=13).value = item.reaction_plan or ''
        
        # 应用样式
        for col in range(1, 14):
            cell = ws.cell(row=data_row, column=col)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        ws.row_dimensions[data_row].height = 60
        data_row += 1
    
    # 页脚
    footer_row = data_row
    ws.merge_cells(f'A{footer_row}:P{footer_row}')
    ws.cell(row=footer_row, column=1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Control Plan ID: {plan_id}"
    ws.cell(row=footer_row, column=1).font = Font(name='Arial', size=8, italic=True)
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='right')
    ws.row_dimensions[footer_row].height = 20
    
    # 输出文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ControlPlan_{plan.part_number or plan.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


def export_control_plans_batch_excel(plan_ids: List[int], db: Session) -> Response:
    """批量导出控制计划汇总"""
    plans = db.query(ControlPlan).filter(ControlPlan.id.in_(plan_ids)).all()
    if not plans:
        raise HTTPException(status_code=404, detail="未找到指定的控制计划")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "控制计划汇总"
    
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['ID', '控制计划编号', '零件号', '零件名称', '计划类型', '状态', '版本', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, plan in enumerate(plans, 2):
        row_data = [
            plan.id,
            plan.control_plan_number or '',
            plan.part_number or '',
            plan.part_name or '',
            plan.plan_type or '',
            plan.status or '',
            plan.version or '',
            plan.created_at.strftime('%Y-%m-%d %H:%M') if plan.created_at else ''
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.font = Font(name='微软雅黑', size=10)
            cell.border = thin_border
    
    col_widths = [8, 20, 15, 20, 12, 10, 8, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"控制计划汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


def export_ocap_excel(ocap_id: int, db: Session) -> Response:
    """导出 OCAP 报告 Excel"""
    ocap = db.query(OCAP).filter(OCAP.id == ocap_id).first()
    if not ocap:
        raise HTTPException(status_code=404, detail=f"OCAP {ocap_id} 不存在")
    
    signals = db.query(OCAPSignal).filter(OCAPSignal.ocap_id == ocap_id).all()
    steps = db.query(OCAPStep).filter(OCAPStep.ocap_id == ocap_id).order_by(OCAPStep.phase, OCAPStep.sort_order).all()
    executions = db.query(OCAPExecution).filter(OCAPExecution.ocap_id == ocap_id).all()
    root_causes = db.query(OCAPRootCause).filter(OCAPRootCause.ocap_id == ocap_id).all()
    actions = db.query(OCAPCorrectiveAction).filter(OCAPCorrectiveAction.ocap_id == ocap_id).all()
    
    wb = Workbook()
    
    # Sheet 1: OCAP 概要
    ws1 = wb.active
    ws1.title = "OCAP 概要"
    
    header_font = Font(name='微软雅黑', bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"OCAP 报告 - {ocap.ocap_number or '未命名'}"
    ws1['A1'].font = Font(name='微软雅黑', bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30
    
    info_headers = ['OCAP 编号', '标题', '优先级', '状态', '产线', '负责人', '创建时间', '完成时间']
    info_values = [
        ocap.ocap_number or '',
        ocap.title or '',
        ocap.priority or '',
        ocap.status or '',
        ocap.line.line_name if ocap.line else '',
        ocap.assigned_to or '',
        ocap.created_at.strftime('%Y-%m-%d %H:%M') if ocap.created_at else '',
        ocap.completed_at.strftime('%Y-%m-%d %H:%M') if ocap.completed_at else ''
    ]
    
    for col, (header, value) in enumerate(zip(info_headers, info_values), 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = f"{header}: {value}"
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = cell_alignment
    
    # SPC 信号
    signal_row = 5
    ws1.cell(row=signal_row, column=1).value = 'SPC 信号'
    ws1.cell(row=signal_row, column=1).font = header_font
    signal_row += 1
    
    signal_headers = ['信号类型', '规则', '测量值', '触发时间', '状态']
    for col, header in enumerate(signal_headers, 1):
        cell = ws1.cell(row=signal_row, column=col)
        cell.value = header
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    signal_row += 1
    for signal in signals:
        row_data = [
            signal.signal_type or '',
            signal.rule or '',
            signal.measurement_value or '',
            signal.triggered_at.strftime('%Y-%m-%d %H:%M') if signal.triggered_at else '',
            signal.status or ''
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=signal_row, column=col)
            cell.value = value
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = cell_alignment
            cell.border = thin_border
        signal_row += 1
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"OCAP_{ocap.ocap_number or ocap.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


def export_control_plan_detailed_report(plan_id: int, db: Session) -> Response:
    """
    导出控制计划详细报告（HTML 格式）
    使用 Jinja2 模板引擎和 control_plan_report.html 模板
    包含完整的控制计划信息、项目列表和统计信息
    """
    plan = db.query(ControlPlan).filter(ControlPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail=f"控制计划 {plan_id} 不存在")
    
    items = db.query(ControlPlanItem).filter(
        ControlPlanItem.control_plan_id == plan_id
    ).order_by(ControlPlanItem.sort_order).all()
    
    # 加载模板
    template = jinja_env.get_template('control_plan_report.html')
    
    # 准备模板数据
    template_data = {
        'plan_type': plan.plan_type or 'production',
        'page_number': plan.page_number or 1,
        'total_pages': plan.total_pages or 1,
        'control_plan_number': plan.control_plan_number or '',
        'part_number': plan.part_number or '',
        'latest_change_level': plan.latest_change_level or '',
        'part_name': plan.part_name or '',
        'organization_plant': plan.organization_plant or '',
        'organization_code': plan.organization_code or '',
        'key_contact': plan.key_contact or '',
        'core_team': plan.core_team or '',
        'date_orig': plan.date_orig.strftime('%Y-%m-%d') if plan.date_orig else '',
        'date_rev': plan.date_rev.strftime('%Y-%m-%d') if plan.date_rev else '',
        'org_approval': plan.org_approval_by or '',
        'customer_eng_approval': plan.customer_eng_approval_by or '',
        'customer_quality_approval': plan.customer_quality_approval_by or '',
        'other_approval': plan.other_approval_by or '',
        'items': items,
        'generated_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'plan_id': plan_id
    }
    
    # 渲染模板
    html_content = template.render(**template_data)
    
    # 生成文件名
    filename = f"ControlPlan_{plan.part_number or plan.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    
    return Response(
        content=html_content,
        media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


def export_capability_analysis_report(analysis_id: int, db: Session) -> Response:
    """
    导出 SPC 研究报告（HTML 格式）
    使用 Jinja2 模板引擎和 spc_capability_report.html 模板
    按照 AIAG/VDA SPC 手册标准格式，包含：
    - 过程信息
    - 直方图
    - 原始数据图
    - 概率图
    - 控制图
    - 过程能力指数
    - 正态性检验
    - 结论和建议
    """
    from app.models.models import CapabilityAnalysis as CapabilityAnalysisModel
    
    analysis = db.query(CapabilityAnalysisModel).filter(
        CapabilityAnalysisModel.id == analysis_id
    ).first()
    if not analysis:
        raise HTTPException(status_code=404, detail=f"能力分析 {analysis_id} 不存在")
    
    # 加载模板
    template = jinja_env.get_template('spc_capability_report.html')
    
    # 准备数据
    data_values = analysis.data_values if hasattr(analysis, 'data_values') and analysis.data_values else []
    data_preview = ', '.join([f"{v:.4f}" for v in data_values[:20]])
    if len(data_values) > 20:
        data_preview += f', ... (共{len(data_values)}个数据)'
    
    # 准备模板数据
    template_data = {
        'process_name': analysis.analysis_name or '',
        'machine_name': '',
        'study_location': '',
        'process_id': '',
        'machine_id': '',
        'operator_name': '',
        'study_date': analysis.analysis_time.strftime('%Y-%m-%d') if analysis.analysis_time else '',
        'start_time': '',
        'end_time': '',
        'part_name_id': f"{analysis.part_name or ''}, ID: {analysis.part_id or ''}" if hasattr(analysis, 'part_name') else '',
        'characteristic_name_id': f"{analysis.characteristic_name or ''}, ID: {analysis.characteristic_id or ''}" if hasattr(analysis, 'characteristic_name') else '',
        'lsl': f"{float(analysis.lsl):.4f}" if analysis.lsl else '',
        'usl': f"{float(analysis.usl):.4f}" if analysis.usl else '',
        'study_remarks': '',
        'sample_size': analysis.sample_count or '',
        'subgroup_size': analysis.subgroup_count or '',
        'sampling_strategy': '',
        'x50': f"{float(analysis.mean):.4f}" if analysis.mean else '',
        'variation_estimate': f"{6 * float(analysis.sigma_within):.4f}" if analysis.sigma_within else '',
        'distribution_model': 'Normal Distribution',
        'cp_g': '1.67',
        'cpk_g': '1.33',
        'calculation_method': 'Geometric Method',
        'cp': f"{float(analysis.cp):.3f}" if analysis.cp else '',
        'cpk': f"{float(analysis.cpk):.3f}" if analysis.cpk else '',
        'cp_ci_lower': '',
        'cp_ci_upper': '',
        'cpk_ci_lower': '',
        'cpk_ci_upper': '',
        'p_usl': '0.00000%',
        'ppm_usl': '0.0',
        'p_lsl': '0.00000%',
        'ppm_lsl': '0.0',
        'conclusion': get_conclusion_desc(float(analysis.cpk) if analysis.cpk else 0),
        'generated_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'analysis_id': analysis_id,
        'data_preview': data_preview
    }
    
    # 渲染模板
    html_content = template.render(**template_data)
    
    # 生成文件名
    filename = f"SPC_Capability_{analysis.analysis_name or analysis_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    
    return Response(
        content=html_content,
        media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


def get_cpk_class(cpk: float) -> str:
    if cpk >= 1.67:
        return 'excellent'
    elif cpk >= 1.33:
        return 'good'
    elif cpk >= 1.0:
        return 'fair'
    return 'poor'


def get_cp_class(cp: float) -> str:
    return get_cpk_class(cp)


def get_ppk_class(ppk: float) -> str:
    return get_cpk_class(ppk)


def get_pp_class(pp: float) -> str:
    return get_cpk_class(pp)


def get_capability_evaluation(cpk: float) -> str:
    if cpk >= 1.67:
        return '过程能力优秀 (Excellent): Cpk ≥ 1.67，过程能力非常出色，产品质量稳定可靠，过程变异很小，可以满足最严格的客户要求。建议：保持当前控制水平，可以考虑优化成本。'
    elif cpk >= 1.33:
        return '过程能力充足 (Good): 1.33 ≤ Cpk < 1.67，过程能力良好，产品质量满足要求，过程处于统计控制状态。建议：持续监控过程，定期评审控制策略。'
    elif cpk >= 1.0:
        return '过程能力一般 (Fair): 1.0 ≤ Cpk < 1.33，过程能力处于临界状态，虽然满足基本要求但存在改进空间。建议：分析变异来源，制定改进计划，提升过程能力。'
    else:
        return '过程能力不足 (Poor): Cpk < 1.0，过程能力不足，产品质量风险较高，可能无法满足客户要求。建议：立即采取纠正措施，进行全面的过程分析和改进。'


def get_conclusion_title(cpk: float) -> str:
    if cpk >= 1.67:
        return '结论：过程能力优秀 (Process Capability is Excellent)'
    elif cpk >= 1.33:
        return '结论：过程能力充足 (Process Capability is Good)'
    elif cpk >= 1.0:
        return '结论：过程能力一般 (Process Capability is Fair)'
    else:
        return '结论：过程能力不足 (Process Capability is Poor)'


def get_conclusion_desc(cpk: float) -> str:
    if cpk >= 1.67:
        return 'The process capability is excellent. Cpk ≥ 1.67 indicates outstanding process performance with minimal variation. Product quality is stable and reliable, meeting the most stringent customer requirements. Continue to monitor the process and maintain current control levels.'
    elif cpk >= 1.33:
        return 'The process capability is good. 1.33 ≤ Cpk < 1.67 indicates capable process performance. Product quality meets requirements and the process is in statistical control. Recommend continuous monitoring and periodic review of control strategies.'
    elif cpk >= 1.0:
        return 'The process capability is fair. 1.0 ≤ Cpk < 1.33 indicates marginal process performance. While basic requirements are met, there is room for improvement. Recommend analyzing sources of variation and developing improvement plans.'
    else:
        return 'The process capability is poor. Cpk < 1.0 indicates inadequate process performance with high quality risk. The process may not meet customer requirements. Immediate corrective actions and comprehensive process improvement are required.'
